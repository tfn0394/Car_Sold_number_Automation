from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import requests
import pandas as pd
from bs4 import BeautifulSoup
import time
import re
import openpyxl
from openpyxl import Workbook, load_workbook
import string
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl as xl;
from colorama import Back, Fore, Style
from openpyxl.utils import cell
import calendar
import arrow
from datetime import date, datetime, timedelta
import csv
import ftplib
from getpass import getpass
import os
import glob
import math
from pathlib import Path
from paramiko import *
import myconstants
from openpyxl.styles import Alignment

PATH = "C:\Program Files (x86)\chromedriver.exe"

driver = webdriver.Chrome(PATH)

driver.get("https://qap.nzta.govt.nz/single/?appid=2e5bd26c-5142-485e-a96d-ce0d903b3b5b&sheet=22c77866-93e2-4632-a65f-7b171ae93647&opt=ctxmenu")

driver.refresh()

driver.refresh()

time.sleep(4)

print (Fore.GREEN + '\n------------------------------------------------------------------------------------------------------\n')
print("                                          Data Extracting ...\n")
print('------------------------------------------------------------------------------------------------------')

time.sleep(3)

try:
    #------------------------------------------------------------------Click Import status on Top
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "PcL_content"))
    )
    element.click()

    time.sleep(2)

    #------------------------------------------------------------------Click New on list
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH,"//ul[@class='qv-listbox ng-scope']/li[1]"))
    )
    element.click()

    time.sleep(2)

    #------------------------------------------------------------------Click Vechicle Type on Top
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "ZDLC_content"))
    )
    element.click()

    time.sleep(3)

    #------------------------------------------------------------------Click 5 buttons on list
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH,"//ul[@class='qv-listbox ng-scope']/li[1]"))
    )
    element.click()

    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH,"//ul[@class='qv-listbox ng-scope']/li[2]"))
    )
    element.click()

    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH,"//ul[@class='qv-listbox ng-scope']/li[3]"))
    )
    element.click()

    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH,"//ul[@class='qv-listbox ng-scope']/li[4]"))
    )
    element.click()

    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH,"//ul[@class='qv-listbox ng-scope']/li[5]"))
    )
    element.click()

    time.sleep(3)

    #------------------------------------------------------------------Search box location
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/div[5]/div/div/div/ng-transclude/div/div[3]/div/article/div[1]/div/div/div/div[1]/div/input"))
    )

    #------------------------------------------------------------------Fill value in Search box
    element.send_keys("PASSENGER CAR/VAN")

    time.sleep(3)

    #------------------------------------------------------------------Click first button on list
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH,"//ul[@class='qv-listbox ng-scope']/li[1]"))
    )
    element.click()

    time.sleep(2)

    #------------------------------------------------------------------Clear Search box
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH,"/html/body/div[5]/div/div/div/ng-transclude/div/div[3]/div/article/div[1]/div/div/div/div[1]/div/input"))
    )
    element.clear()

    time.sleep(2)

    #------------------------------------------------------------------Fill value in Search box
    element.send_keys("SPECIAL PURPOSE VEHICLE")

    time.sleep(2)

    #------------------------------------------------------------------Click first button on list
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH,"//ul[@class='qv-listbox ng-scope']/li[1]"))
    )
    element.click()

    time.sleep(2)

    #------------------------------------------------------------------Click Vehicle type on Side
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, "//div[@class='sfs_lui sfYtvqAhe_8a56794b-7d3b-40cf-98de-964dabb14341']/label[@title='Vehicle type']"))
    )
    element.click()

    time.sleep(2)

    #------------------------------------------------------------------Table location
    tables = WebDriverWait(driver, 20).until(
        EC.presence_of_all_elements_located((By.TAG_NAME, "table")))

    #------------------------------------------------------------------Find passenger and commercial number
    for table in tables:
        df = pd.read_html(table.get_attribute('outerHTML'))
        df = df[0].replace(to_replace=[r"[-()]", "\s+"], value=["", ""], regex=True)
        p_number = df.iloc[15,4]
        c_number = df.iloc[0,4] + df.iloc[3,4] + df.iloc[6,4] + df.iloc[9,4] + df.iloc[12,4] + df.iloc[18,4]

    print(Style.RESET_ALL)
    print('                                              Sheet1\n')
    print('------------------------------------------------------------------------------------------------------\n')

    print('New Zealand New vehicle sales of Passenger Registration count this month:', p_number,'\n')
    print('------------------------------------------------------------------------------------------------------\n')

    print('New Zealand New vehicle sales of Commercial Registration count this month:', c_number,'\n')
    print('------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Load NewVehicleSales-NZ excel file
    workbook = openpyxl.load_workbook("K:/interest-nz/Statistics NZ/NewVehicleSales-NZ.xlsx")
    #workbook = openpyxl.load_workbook("C:/Users/User/Desktop/Python_Excel/NewVehicleSales-NZ-copy.xlsx")
 
    #------------------------------------------------------------------Open NewVehicleSales-NZ sheet 1
    sheet1 = workbook['Vehicles']

    #------------------------------------------------------------------Get the last day of the previous month
    today = date.today()
    first_day_of_month = date(today.year, today.month, 1)
    last_day_of_previous_month = first_day_of_month - timedelta(days=1)

    #------------------------------------------------------------------Get the last row with data in column B
    print('The last day of previous month:', last_day_of_previous_month,'\n')
    print('------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Get the last row with data in column B
    last_row = sheet1.max_row

    while sheet1.cell(row=last_row, column=2).value is None and last_row > 1:
        last_row -= 1

    # Create an Alignment object to align the data to the right
    alignment_right = Alignment(horizontal='right')

    # Apply the alignment to the cell
    sheet1.cell(row=last_row + 1, column=2).alignment = alignment_right

    sheet1.cell(row=last_row + 1, column=2).value = last_day_of_previous_month.strftime("%b-%y")

    #------------------------------------------------------------------Get the last row with data in column C
    last_row = sheet1.max_row

    while sheet1.cell(row=last_row, column=3).value is None and last_row > 1:
        last_row -= 1

    sheet1.cell(row=last_row + 1, column=3).value = p_number

    #------------------------------------------------------------------Get the last row with data in column D
    last_row = sheet1.max_row

    while sheet1.cell(row=last_row, column=4).value is None and last_row > 1:
        last_row -= 1

    sheet1.cell(row=last_row + 1, column=4).value = c_number

    #------------------------------------------------------------------Double Click Import status on Top
    time.sleep(2)

    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "PcL_content"))
    )
    element.click()

    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "PcL_content"))
    )
    element.click()

    time.sleep(3)

    #------------------------------------------------------------------Click New on list
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH,"//ul[@class='qv-listbox ng-scope']/li[1]"))
    )
    element.click()

    time.sleep(3)

    #------------------------------------------------------------------Click Used on list
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH,"//ul[@class='qv-listbox ng-scope']/li[4]"))
    )
    element.click()

    time.sleep(2)

    #------------------------------------------------------------------Table location
    tables = WebDriverWait(driver, 20).until(
        EC.presence_of_all_elements_located((By.TAG_NAME, "table")))
    
    time.sleep(2)

    #------------------------------------------------------------------Find passenger used number
    for table in tables:
        df = pd.read_html(table.get_attribute('outerHTML'))
        df = df[0].replace(to_replace=[r"[-()]", "\s+"], value=["", ""], regex=True)
        up_number = df.iloc[15,4]    

    time.sleep(3)

    print('New Zealand Used vehicle sales of Passenger Registration count this month:', up_number,'\n')
    print('------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and print last cell of S
    last_row = sheet1.max_row

    while sheet1.cell(row=last_row, column=19).value is None and last_row > 1:
        last_row -= 1

    sheet1.cell(row=last_row + 1, column=19).value = up_number

    time.sleep(2)

    #------------------------------------------------------------------Modify and Calculate last cell of E
    print('Last value of E: ', p_number + c_number)

    last_row = sheet1.max_row

    while sheet1.cell(row=last_row, column=5).value is None and last_row > 1:
        last_row -= 1

    sheet1.cell(row=last_row + 1, column=5).value = p_number + c_number

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of G
    # Find the last 13th cell in column C that contains value
    cell_count = 0

    for cell in reversed(sheet1['C']):
        if cell.value is not None:
            cell_count += 1
            if cell_count == 13:
                last_13th_cell_C = cell
                break

    print('Last 13th value in C:', last_13th_cell_C.value)

    print('Last value of G: ', ((p_number/last_13th_cell_C.value)-1)*100)

    last_row = sheet1.max_row

    while sheet1.cell(row=last_row, column=7).value is None and last_row > 1:
        last_row -= 1

    sheet1.cell(row=last_row + 1, column=7).value = ((p_number/last_13th_cell_C.value)-1)*100

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of H
    # Find the last 13th cell in column D that contains value
    cell_count = 0

    for cell in reversed(sheet1['D']):
        if cell.value is not None:
            cell_count += 1
            if cell_count == 13:
                last_13th_cell_D = cell
                break

    print('Last 13th value in D:', last_13th_cell_D.value)

    print('Last value of H: ', ((c_number/last_13th_cell_D.value)-1)*100)

    last_row = sheet1.max_row

    while sheet1.cell(row=last_row, column=8).value is None and last_row > 1:
        last_row -= 1

    sheet1.cell(row=last_row + 1, column=8).value = ((c_number/last_13th_cell_D.value)-1)*100

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of I
    # Find the last 13th cell in column E that contains value
    cell_count = 0

    for cell in reversed(sheet1['E']):
        if cell.value is not None:
            cell_count += 1
            if cell_count == 13:
                last_13th_cell_E = cell
                break

    print('Last 13th value in E:', last_13th_cell_E.value)
    
    print('Last value of I: ', (((p_number + c_number)/last_13th_cell_E.value)-1)*100)

    last_row = sheet1.max_row

    while sheet1.cell(row=last_row, column=9).value is None and last_row > 1:
        last_row -= 1

    sheet1.cell(row=last_row + 1, column=9).value = (((p_number + c_number)/last_13th_cell_E.value)-1)*100

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of K
    # Find the last row with a value in column C

    last_row = sheet1.max_row

    while sheet1.cell(row=last_row, column=3).value is None and last_row > 1:
        last_row -= 1

    # Calculate the sum of the last cells in column C that contain a value
    sum_range_C = sheet1['C' + str(last_row-11):'C' + str(last_row)]
    sum_value_K = sum(cell.value for row in sum_range_C for cell in row)

    print('Last value of K: ', sum_value_K)

    sheet1.cell(row=last_row, column=11).value = sum_value_K #114401


    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of L
    # Find the last row with a value in column D

    last_row = sheet1.max_row

    while sheet1.cell(row=last_row, column=4).value is None and last_row > 1:
        last_row -= 1

    # Calculate the sum of the last cells in column D that contain a value
    sum_range_D = sheet1['D' + str(last_row-11):'D' + str(last_row)]
    sum_value_L = sum(cell.value for row in sum_range_D for cell in row)

    print('Last value of L: ', sum_value_L)

    sheet1.cell(row=last_row, column=12).value = sum_value_L

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of M
    # Find the last row with a value in column E

    last_row = sheet1.max_row

    while sheet1.cell(row=last_row, column=5).value is None and last_row > 1:
        last_row -= 1

    # Calculate the sum of the last cells in column E that contain a value
    sum_range_E = sheet1['E' + str(last_row-11):'E' + str(last_row)]
    sum_value_M = sum(cell.value for row in sum_range_E for cell in row)

    print('Last value of M: ',sum_value_M)

    sheet1.cell(row=last_row, column=13).value = sum_value_M

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of O
    # Find the last 13th cell in column K that contains value
    cell_count = 0

    for cell in reversed(sheet1['K']):
        if cell.value is not None:
            cell_count += 1
            if cell_count == 13:
                last_13th_cell_K = cell
                break

    print('Last 13th value in K:', last_13th_cell_K.value)

    print('Last value of O: ', ((sum_value_K/last_13th_cell_K.value)-1)*100)

    last_row = sheet1.max_row

    while sheet1.cell(row=last_row, column=15).value is None and last_row > 1:
        last_row -= 1

    sheet1.cell(row=last_row + 1, column=15).value = ((sum_value_K/last_13th_cell_K.value)-1)*100

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of P
    # Find the last 13th cell in column L that contains value
    cell_count = 0

    for cell in reversed(sheet1['L']):
        if cell.value is not None:
            cell_count += 1
            if cell_count == 13:
                last_13th_cell_L = cell
                break

    print('Last 13th value in L:', last_13th_cell_L.value)

    print('Last value of P: ', ((sum_value_L/last_13th_cell_L.value)-1)*100)

    last_row = sheet1.max_row

    while sheet1.cell(row=last_row, column=16).value is None and last_row > 1:
        last_row -= 1

    sheet1.cell(row=last_row + 1, column=16).value = ((sum_value_L/last_13th_cell_L.value)-1)*100

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of Q
    # Find the last 13th cell in column M that contains value
    cell_count = 0

    for cell in reversed(sheet1['M']):
        if cell.value is not None:
            cell_count += 1
            if cell_count == 13:
                last_13th_cell_M = cell
                break

    print('Last 13th value in M:', last_13th_cell_M.value)

    print('Last value of Q: ', ((sum_value_M/last_13th_cell_M.value)-1)*100)

    last_row = sheet1.max_row

    while sheet1.cell(row=last_row, column=17).value is None and last_row > 1:
        last_row -= 1

    sheet1.cell(row=last_row + 1, column=17).value = ((sum_value_M/last_13th_cell_M.value)-1)*100

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of R
    print('Last value of R: ', p_number + up_number)

    last_row = sheet1.max_row

    while sheet1.cell(row=last_row, column=18).value is None and last_row > 1:
        last_row -= 1

    sheet1.cell(row=last_row + 1, column=18).value = p_number + up_number

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of T
    # Find the last 13th cell in column S that contains value
    cell_count = 0

    for cell in reversed(sheet1['S']):
        if cell.value is not None:
            cell_count += 1
            if cell_count == 13:
                last_13th_cell_S = cell
                break

    print('Last 13th value in S:', last_13th_cell_S.value)

    print('Last value of T: ', ((up_number/last_13th_cell_S.value)-1)*100)

    last_row = sheet1.max_row

    while sheet1.cell(row=last_row, column=20).value is None and last_row > 1:
        last_row -= 1

    sheet1.cell(row=last_row + 1, column=20).value = ((up_number/last_13th_cell_S.value)-1)*100

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of U
    # Find the last row with a value in column S

    last_row = sheet1.max_row

    while sheet1.cell(row=last_row, column=19).value is None and last_row > 1:
        last_row -= 1

    # Calculate the sum of the last cells in column E that contain a value
    sum_range_S = sheet1['S' + str(last_row-11):'S' + str(last_row)]
    sum_value_U = sum(cell.value for row in sum_range_S for cell in row)

    print('Last value of U: ', sum_value_U)

    sheet1.cell(row=last_row, column=21).value = sum_value_U

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of V
    print('Last value of V: ', sum_value_M + sum_value_U)

    last_row = sheet1.max_row

    while sheet1.cell(row=last_row, column=22).value is None and last_row > 1:
        last_row -= 1

    sheet1.cell(row=last_row + 1, column=22).value = sum_value_M + sum_value_U

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of W
    print('Last value of W: ', p_number + c_number + up_number)

    last_row = sheet1.max_row

    while sheet1.cell(row=last_row, column=23).value is None and last_row > 1:
        last_row -= 1

    sheet1.cell(row=last_row + 1, column=23).value = p_number + c_number + up_number

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of Y
    print('Last value of Y: ', ((p_number/last_13th_cell_C.value)-1)*100)

    last_row = sheet1.max_row

    while sheet1.cell(row=last_row, column=25).value is None and last_row > 1:
        last_row -= 1

    sheet1.cell(row=last_row + 1, column=25).value = ((p_number/last_13th_cell_C.value)-1)*100

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of AA
    # Find the last 13th cell in column U that contains value
    cell_count = 0

    for cell in reversed(sheet1['U']):
        if cell.value is not None:
            cell_count += 1
            if cell_count == 13:
                last_13th_cell_U = cell
                break

    print('Last 13th value in U:', last_13th_cell_U.value)

    print('Last value of AA: ', ((sum_value_U/last_13th_cell_U.value)-1)*100)

    last_row = sheet1.max_row

    while sheet1.cell(row=last_row, column=27).value is None and last_row > 1:
        last_row -= 1

    sheet1.cell(row=last_row + 1, column=27).value = ((sum_value_U/last_13th_cell_U.value)-1)*100

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of A
    print('Last value of A: ', (sum_value_L/sum_value_M)*100,'%')

    last_row = sheet1.max_row

    while sheet1.cell(row=last_row, column=1).value is None and last_row > 1:
        last_row -= 1

    sheet1.cell(row=last_row + 1, column=1).value = sum_value_L/sum_value_M

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Click Clear button on right
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "5fea0419-0546-4005-8dbb-94d300c1fa65_content"))
    )
    element.click()

    #------------------------------------------------------------------Click Month year on Top
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "ThaDGPu_content"))
    )
    element.click()

    time.sleep(2)

    #------------------------------------------------------------------Search box location
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/div[5]/div/div/div/ng-transclude/div/div[3]/div/article/div[1]/div/div/div/div[1]/div/input"))
    )

    #------------------------------------------------------------------Fill value in Search box
    # Get the current date
    now = arrow.now()

    # Get the name of the previous month (e.g. "Apr")
    prev_month_name = now.shift(months=-1).format("MMM")

    # Get the current year (e.g. "2023")
    current_year = datetime.now().strftime("%Y")

    element.send_keys(prev_month_name + " " + current_year)

    time.sleep(2)

    #------------------------------------------------------------------Click first button on list
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH,"//ul[@class='qv-listbox ng-scope']/li[1]"))
    )
    element.click()

    time.sleep(2)

    #------------------------------------------------------------------Click Import status on Top
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "PcL_content"))
    )
    element.click()

    time.sleep(2)

    #------------------------------------------------------------------Click New on list
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH,"//ul[@class='qv-listbox ng-scope']/li[1]"))
    )
    element.click()

    time.sleep(2)

    #------------------------------------------------------------------Click Vechicle Type on Top
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.ID, "ZDLC_content"))
    )
    element.click()

    time.sleep(2)

    #------------------------------------------------------------------Search box location
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/div[5]/div/div/div/ng-transclude/div/div[3]/div/article/div[1]/div/div/div/div[1]/div/input"))
    )

    #------------------------------------------------------------------Fill value in Search box
    element.send_keys("PASSENGER CAR/VAN")

    time.sleep(2)

    #------------------------------------------------------------------Click first button on list
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH,"//ul[@class='qv-listbox ng-scope']/li[1]"))
    )
    element.click()

    time.sleep(2)

    #------------------------------------------------------------------Click Motive power on Side
    element = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, "//div[@class='sfs_lui sfYtvqAhe_8a56794b-7d3b-40cf-98de-964dabb14341']/label[@title='Motive power']"))
    )
    element.click()

    time.sleep(2)

    #------------------------------------------------------------------Table location
    tables = WebDriverWait(driver, 20).until(
        EC.presence_of_all_elements_located((By.TAG_NAME, "table")))

    #------------------------------------------------------------------Find electric number
    for table in tables:
        df = pd.read_html(table.get_attribute('outerHTML'))
        df = df[0].replace(to_replace=[r"[-()]", "\s+"], value=["", ""], regex=True)
        e_number = df.iloc[6,4]
        h_number = df.iloc[9,4]
        t_nev_number = e_number + h_number + df.iloc[18,4] + df.iloc[21,4]

    print('                                              Sheet3\n')
    print('------------------------------------------------------------------------------------------------------\n')

    print('New Zealand New Electric vehicle sales of Passenger Registration count this month:', e_number,'\n')
    print('------------------------------------------------------------------------------------------------------\n')

    print('New Zealand New Hydrogen vehicle sales of Passenger Registration count this month:', h_number,'\n')
    print('------------------------------------------------------------------------------------------------------\n')

    print('New Zealand New NEV vehicle sales of Passenger Registration count this month:', t_nev_number,'\n')
    print('------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Open NewVehicleSales-NZ sheet 2
    sheet3 = workbook['NEVs']

    #------------------------------------------------------------------Get the last day of the previous month
    today = date.today()
    first_day_of_month = date(today.year, today.month, 1)
    last_day_of_previous_month = first_day_of_month - timedelta(days=1)

    #------------------------------------------------------------------Get the last row with data in column B
    print('The last day of previous month:', last_day_of_previous_month,'\n')
    print('------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Get the last row with data in column B
    last_row = sheet3.max_row

    while sheet3.cell(row=last_row, column=2).value is None and last_row > 1:
        last_row -= 1

    # Apply the alignment to the cell
    sheet3.cell(row=last_row + 1, column=2).alignment = alignment_right

    sheet3.cell(row=last_row + 1, column=2).value = last_day_of_previous_month.strftime("%b-%y")

    #------------------------------------------------------------------Modify last cell of C
    last_row = sheet3.max_row

    while sheet3.cell(row=last_row, column=3).value is None and last_row > 1:
        last_row -= 1

    sheet3.cell(row=last_row + 1, column=3).value = e_number

    #------------------------------------------------------------------Modify last cell of F
    last_row = sheet3.max_row

    while sheet3.cell(row=last_row, column=6).value is None and last_row > 1:
        last_row -= 1

    sheet3.cell(row=last_row + 1, column=6).value = h_number

    #------------------------------------------------------------------Modify last cell of L
    last_row = sheet3.max_row

    while sheet3.cell(row=last_row, column=12).value is None and last_row > 1:
        last_row -= 1

    sheet3.cell(row=last_row + 1, column=12).value = t_nev_number

    #------------------------------------------------------------------Modify and print last cell of M
    print('Last value of M: ', t_nev_number/p_number)

    last_row = sheet3.max_row

    while sheet3.cell(row=last_row, column=13).value is None and last_row > 1:
        last_row -= 1

    # Apply the alignment to the cell
    sheet3.cell(row=last_row + 1, column=13).alignment = alignment_right

    sheet3.cell(row=last_row + 1, column=13).value = (format(t_nev_number/p_number, ",.1%"))

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of N
    # Find the last row with a value in column L
    last_row = sheet3.max_row

    while sheet3.cell(row=last_row, column=12).value is None and last_row > 1:
        last_row -= 1

    # Calculate the sum of the last cells in column L that contain a value and divide by passenger number
    sum_range_NL = sheet3['L' + str(last_row-11):'L' + str(last_row)]
    sum_value_NL = sum(cell.value for row in sum_range_NL for cell in row)

    print('Last value of N: ', sum_value_NL/sum_value_K)

    # Apply the alignment to the cell
    sheet3.cell(row=last_row, column=14).alignment = alignment_right

    sheet3.cell(row=last_row, column=14).value = (format(sum_value_NL/sum_value_K, ",.1%"))

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of P
    print('Last value of P: ', sum_value_NL)

    last_row = sheet3.max_row

    while sheet3.cell(row=last_row, column=16).value is None and last_row > 1:
        last_row -= 1

    sheet3.cell(row=last_row + 1, column=16).value = sum_value_NL

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of Q

    last_row = sheet3.max_row

    while sheet3.cell(row=last_row, column=3).value is None and last_row > 1:
        last_row -= 1

    # Calculate the sum of the last cells in column C that contain a value
    sum_range_NC = sheet3['C' + str(last_row-11):'C' + str(last_row)]
    sum_value_NC = sum(cell.value for row in sum_range_NC for cell in row)

    print('Last value of Q: ', sum_value_NC)

    sheet3.cell(row=last_row, column=17).value = sum_value_NC

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of R

    print('Last value of R: ', sum_value_NC/sum_value_NL)

    last_row = sheet3.max_row

    while sheet3.cell(row=last_row, column=18).value is None and last_row > 1:
        last_row -= 1

    sheet3.cell(row=last_row + 1, column=18).value = sum_value_NC/sum_value_NL

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of T

    print('Last value of T: ', e_number/t_nev_number)

    last_row = sheet3.max_row

    while sheet3.cell(row=last_row, column=20).value is None and last_row > 1:
        last_row -= 1

    sheet3.cell(row=last_row + 1, column=20).value = e_number/t_nev_number

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Modify and Calculate last cell of U

    # Find the sum of last 13th value in column C
    last_row = sheet3.max_row

    while sheet3.cell(row=last_row, column=3).value is None and last_row > 1:
        last_row -= 1

    # Calculate the sum of the last cells in column C that contain a value
    sum_range_NC13 = sheet3['C' + str(last_row-12):'C' + str(last_row)]
    sum_value_NC13 = sum(cell.value for row in sum_range_NC13 for cell in row) 

    # Find the sum of last 13th value in column L
    
    last_row = sheet3.max_row

    while sheet3.cell(row=last_row, column=12).value is None and last_row > 1:
        last_row -= 1

    # Calculate the sum of the last cells in column C that contain a value
    sum_range_NL13 = sheet3['L' + str(last_row-12):'L' + str(last_row)]
    sum_value_NL13 = sum(cell.value for row in sum_range_NL13 for cell in row)
    
    print('Last value of U: ', sum_value_NC13/sum_value_NL13)

    last_row = sheet3.max_row

    while sheet3.cell(row=last_row, column=21).value is None and last_row > 1:
        last_row -= 1

    sheet3.cell(row=last_row + 1, column=21).value = sum_value_NC13/sum_value_NL13

    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Update all the csv files
    # Change the date format
    last_day_of_previous_month = datetime.now().replace(day=1) - timedelta(days=1)
    last_day_of_previous_month_str = last_day_of_previous_month.strftime('%d-%b-%y')

    # Dictionary that maps file names or file paths to the corresponding data
    data_dict = {'K:/interest-nz/interest.co.nz/chart_data/industry/newvehicles-cgrowth.csv': [last_day_of_previous_month, ((c_number/last_13th_cell_D.value)-1)*100],
                'K:/interest-nz/interest.co.nz/chart_data/industry/newvehicles-commercial.csv': [last_day_of_previous_month, c_number],
                'K:/interest-nz/interest.co.nz/chart_data/industry/newvehicles-importedcars.csv': [last_day_of_previous_month, up_number],
                'K:/interest-nz/interest.co.nz/chart_data/industry/newvehicles-newcannualsales.csv': [last_day_of_previous_month, sum_value_L],
                'K:/interest-nz/interest.co.nz/chart_data/industry/newvehicles-newpannualsales.csv': [last_day_of_previous_month, sum_value_K],
                'K:/interest-nz/interest.co.nz/chart_data/industry/newvehicles-passenger.csv': [last_day_of_previous_month, p_number],
                'K:/interest-nz/interest.co.nz/chart_data/industry/newvehicles-pgrowth.csv': [last_day_of_previous_month, ((p_number/last_13th_cell_C.value)-1)*100],
                'K:/interest-nz/interest.co.nz/chart_data/industry/newvehicles-usedpannualsales.csv': [last_day_of_previous_month, sum_value_U]}

    # Loop through the data_dict and update each file
    for file_path, new_data in data_dict.items():

        # Read the CSV file into a list of rows
        with open(file_path, mode='r') as csv_file:
            reader = csv.reader(csv_file)
            rows = list(reader)

        # Find the index of the last row with data
        last_row_index = None
        for i in range(len(rows)-1, -1, -1):  # Iterate in reverse order
            if any(rows[i]):  # Check if the row contains any value
                last_row_index = i
                break

        # Insert new data after the last row with data
        rows.insert(last_row_index+1, new_data)

        # Write the updated data back to the CSV file
        with open(file_path, mode='w', newline='') as csv_file:
            writer = csv.writer(csv_file)
            writer.writerows(rows)

    workbook.save(filename="K:/interest-nz/Statistics NZ/NewVehicleSales-NZ.xlsx")
    #workbook.save(filename="C:/Users/User/Desktop/Python_Excel/NewVehicleSales-NZ-copy.xlsx")

    print('                                          Csv files Updated')
    print('\n------------------------------------------------------------------------------------------------------\n')

    #------------------------------------------------------------------Update csv data to website
    host = 'nfs.interest.co.nz'

    transport = Transport(host)
    transport.connect(None, myconstants.USERNAME, myconstants.PASSWORD)
    sftp = SFTPClient.from_transport(transport)

    local_paths = [
    "K:\\interest-nz\\interest.co.nz\\chart_data\\industry\\newvehicles-passenger.csv",
    "K:\\interest-nz\\interest.co.nz\\chart_data\\industry\\newvehicles-usedpannualsales.csv",
    "K:\\interest-nz\\interest.co.nz\\chart_data\\industry\\newvehicles-newpannualsales.csv",
    "K:\\interest-nz\\interest.co.nz\\chart_data\\industry\\newvehicles-importedcars.csv",
    "K:\\interest-nz\\interest.co.nz\\chart_data\\industry\\newvehicles-pgrowth.csv",
    "K:\\interest-nz\\interest.co.nz\\chart_data\\industry\\newvehicles-newcannualsales.csv",
    "K:\\interest-nz\\interest.co.nz\\chart_data\\industry\\newvehicles-commercial.csv",
    "K:\\interest-nz\\interest.co.nz\\chart_data\\industry\\newvehicles-cgrowth.csv"
    ]

    remote_paths = [
    "/var/www/drupal8.interest.co.nz/web/sites/default/files/charts-csv/chart_data/industry/newvehicles-passenger.csv",
    "/var/www/drupal8.interest.co.nz/web/sites/default/files/charts-csv/chart_data/industry/newvehicles-usedpannualsales.csv",
    "/var/www/drupal8.interest.co.nz/web/sites/default/files/charts-csv/chart_data/industry/newvehicles-newpannualsales.csv",
    "/var/www/drupal8.interest.co.nz/web/sites/default/files/charts-csv/chart_data/industry/newvehicles-importedcars.csv",
    "/var/www/drupal8.interest.co.nz/web/sites/default/files/charts-csv/chart_data/industry/newvehicles-pgrowth.csv",
    "/var/www/drupal8.interest.co.nz/web/sites/default/files/charts-csv/chart_data/industry/newvehicles-newcannualsales.csv",
    "/var/www/drupal8.interest.co.nz/web/sites/default/files/charts-csv/chart_data/industry/newvehicles-commercial.csv",
    "/var/www/drupal8.interest.co.nz/web/sites/default/files/charts-csv/chart_data/industry/newvehicles-cgrowth.csv"
    ]

    for i in range(len(local_paths)):
        local_path = local_paths[i]
        remote_path = remote_paths[i]
        if os.path.exists(local_path):
            sftp.put(local_path, remote_path)

    print("                                          Csv files Uploaded")

    print (Fore.GREEN + '\n------------------------------------------------------------------------------------------------------\n')
    print("                                                Done :)\n")
    print('------------------------------------------------------------------------------------------------------\n')

    print(Fore.BLUE + '                                                By Bob')
    print('\n------------------------------------------------------------------------------------------------------')
    print(Style.RESET_ALL)

except:
    driver.quit()
